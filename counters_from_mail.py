from imap_tools import MailBox
from datetime import date, datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import win32com.client
import config, csv, mysql.connector, os

# Initialize the global variable
df = pd.DataFrame()

def readMails():
    global df  # Declare df as global to modify it

    # Initialize a dictionary with model numbers 1 to 28 and default counters set to 0
    data = {model: {'Model': model, 'Color Counter': 0, 'Black Counter': 0, 'Scan Counter': 0} for model in range(1, 29)}

    with MailBox(config.imap_server).login(config.username, config.password, config.counters_dir) as mb:
        for msg in mb.fetch(limit=None, reverse=True, mark_seen=True):
            if str(msg.date.strftime('%Y-%m-%d')) == config.current_day:
                # Split the email content by lines
                mail_msg = msg.text.strip().split('\n')
                # Initialize a temporary dictionary to hold the parsed data
                email_data = {}

                for line in mail_msg:
                    # Split each line by the comma
                    key, value = line.split(',', 1)
                    key = key.strip()
                    value = value.strip()

                    # Parse specific keys and store them in the dictionary
                    if key == '[Model Name]':
                        # Extract the model number and ensure it's between 1 and 28
                        model_number = int(value.split()[0])
                        if 1 <= model_number <= 28:
                            email_data['Model'] = model_number
                    elif key == '[Total Color Counter]':
                        email_data['Color Counter'] = int(value)
                    elif key == '[Total Black Counter]':
                        email_data['Black Counter'] = int(value)
                    elif key == '[Total Scan/Fax Counter]':
                        email_data['Scan Counter'] = int(value)

                # If the model number was found and is valid, update the data dictionary
                if 'Model' in email_data:
                    model_number = email_data['Model']
                    data[model_number] = email_data

    # Convert the data dictionary to a DataFrame
    df = pd.DataFrame.from_dict(data, orient='index')

    # Fill NaN values in 'Scan Counter' with 0
    df['Scan Counter'] = df['Scan Counter'].fillna(0).astype(int)

    # Ensure that all relevant columns are integers
    df = df.astype({
        'Model': 'int32',
        'Color Counter': 'int32',
        'Black Counter': 'int32',
        'Scan Counter': 'int32'
    })

    # Ensure the DataFrame columns are ordered
    df = df[['Model', 'Color Counter', 'Black Counter', 'Scan Counter']]
    df.to_csv(f'{config.currentMonth}_data.csv', index=False)
def dfToMySql():
    global df  # Declare df as global to access it

    # Load the DataFrame from CSV
    df = pd.read_csv(f'{config.currentMonth}_data.csv')

    # Database connection
    conn = mysql.connector.connect(
        host=config.mySQL_host,
        user=config.mySQL_usr,
        password=config.mySQL_pass,
        database=config.mySQL_database  # Ensure correct spelling for 'database'
    )
    cursor = conn.cursor()
    
    # Create a table with counters from the current month
    try:
        cursor.execute(config.createTable)
    except mysql.connector.Error as err:
        print(f"Error: {err}"); cursor.execute(config.dropTable); cursor.execute(config.createTable)

    # Insert latest counters into the current month table
    for row in df.itertuples():
        cursor.execute(f'INSERT INTO {config.currentMonth} (Nr_drukarki, Czarny, Kolor, Skany) VALUES ({row.Model}, {row._3}, {row._2}, {row._4})')

    conn.commit()  # Commit updated values

    # Insert counters from the previous month into the current month table
    cursor.execute(config.insertIntoMonthBefore)
    conn.commit()  # Commit updated values

    # Select all printers with their latest counters and number of printed pages
    cursor.execute(config.query)
    
    headings = ['Nr_drukarki','Czarny','Stary_Czarny','Roznica_Czarny','Kolor','Stary_Kolor','Roznica_Kolor','Skany','Stary_Skany','Roznica_Skany'] # headings
    rows = cursor.fetchall()                                      # read all rows from query

    fp = open(f'.\\{config.currentMonth}.csv', 'w')                         # create .csv file and write to it data from query
    with open(f'.\\{config.currentMonth}.csv', 'w', newline='') as fp:    
        myFile = csv.writer(fp)                                             # Create a CSV writer object
        myFile.writerow(headings)
        for row in rows:                                                # Write the data row by row into the CSV file
            myFile.writerow(row)    
            
    conn.close()
def countersToExcel():
    read_file = pd.read_csv(f'.\\{config.currentMonth}.csv')
    read_file.to_excel(f'.\\{config.currentMonth}.xlsx', index=None, header=True)

    source_filename = f'.\\{config.currentMonth}.xlsx'   
    destination_filename = 'Liczniki.xlsx'
    data_to_copy = pd.read_excel(source_filename)
    
    with pd.ExcelWriter(destination_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        data_to_copy.to_excel(writer, sheet_name=config.currentMonth, index=False)
    
    wb = load_workbook(destination_filename)
    sheet = wb['listaArkusze']
    
    row = 3
    found = False
    while sheet[f"B{row}"].value is not None:
        if sheet[f"B{row}"].value == config.currentMonth:
            found = True
            break
        row += 1

    if found:
        sheet[f"B{row}"] = config.currentMonth  # Replace the existing value if found
    else:
        sheet[f"B{row}"] = config.currentMonth  # Add a new value if not found
    
    # Proceed with updating 'liczniki_do_wyslania' sheet
    sheet = wb['liczniki_do_wyslania']

    formula_text = '=OFFSET(listaArkusze!$B$2,1,0, COUNTA(listaArkusze!$B$2:$B$500),1)'
    dv = DataValidation(type="list", formula1=formula_text, allow_blank=True)
    dv.add(sheet['E3'])
    sheet.add_data_validation(dv)

    # Save the workbook after making changes
    wb.save(destination_filename)

    # Clean up temporary files
    os.remove(f'.\\{config.currentMonth}.csv')
    os.remove(f'.\\{config.currentMonth}.xlsx')
    os.remove(f'.\\{config.currentMonth}_data.csv')
def exportToFinalExcel():
    file_path = f'D:\\Liczniki\\liczniki_{config.currentMonth}.xlsx'
    
    if os.path.isfile(file_path):
        os.remove(file_path)
        
    # Load the source workbook and sheet using openpyxl
    wb = load_workbook('Liczniki.xlsx')
    ws = wb['liczniki_do_wyslania']

    # Update the dropdown cell with the current month
    dropdown_cell = ws['E3']
    dropdown_cell.value = config.currentMonth  # Replace this with the actual current month value

    # Save changes made by openpyxl
    wb.save('Liczniki.xlsx')

    # Open the source workbook using pywin32
    excel = win32com.client.Dispatch("Excel.Application")
    source_wb = excel.Workbooks.Open(r"D:\\Programowanie\\Projects\\Printers_read_from_mail\\Liczniki.xlsx")  # Use the full path
    source_ws = source_wb.Sheets('liczniki_do_wyslania')

    # Create a new workbook
    new_wb = excel.Workbooks.Add()
    new_ws = new_wb.Sheets(1)
    new_ws.Name = f'liczniki_{config.currentMonth}'

    # Copy values (not formulas) to the new sheet
    source_ws.UsedRange.Copy()
    new_ws.Range('A1').PasteSpecial(Paste=-4163)  # -4163 is the constant for pasting values only

    # Save the new workbook
    new_wb.SaveAs(f"D:\Liczniki\\liczniki_{config.currentMonth}")  # Use the full path

    # Close the workbooks
    source_wb.Close(SaveChanges=False)
    new_wb.Close(SaveChanges=True)

    # Quit Excel application
    excel.Quit()

if __name__ == "__main__":
    # Run the functions
    readMails()
    dfToMySql()
    countersToExcel()
    exportToFinalExcel()